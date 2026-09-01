import json
import copy
import os
import re
import shutil
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
import uuid
import zipfile
from datetime import datetime, timezone
from typing import List, Optional
import hashlib

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from .services.pdf_processor import export_ballooned_pdf, page_count, render_page
from .services.gemini_analyzer import analyze_with_gemini, analyze_batch_with_gemini

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROJECT_ROOT = os.path.dirname(BASE)
FRONTEND_DIR = os.path.join(PROJECT_ROOT, 'frontend')
FRONTEND_ASSETS = os.path.join(FRONTEND_DIR, 'assets')

# Runtime storage policy:
# - Local development: backend/data
# - Vercel/serverless: /tmp/ballooning_data (the deployment bundle under /var/task is read-only)
# - Optional override: BALLOONING_DATA_DIR
_DEFAULT_DATA = os.path.join(BASE, 'data')
if os.getenv('VERCEL') or os.path.abspath(os.getcwd()).startswith('/var/task'):
    _DEFAULT_DATA = os.path.join(tempfile.gettempdir(), 'ballooning_data')
DATA = os.path.abspath(os.getenv('BALLOONING_DATA_DIR', _DEFAULT_DATA))
PROJECTS = os.path.join(DATA, 'projects')
LEARNING = os.path.join(DATA, 'learning')
for _folder in (DATA, PROJECTS, LEARNING):
    os.makedirs(_folder, exist_ok=True)

# .env is only a local-development convenience. Vercel values come from Project Environment Variables.
if not os.getenv('VERCEL'):
    load_dotenv(os.path.join(BASE, '.env'))

app = FastAPI(title='InspectBalloon API', version='0.7.0')
app.add_middleware(CORSMiddleware, allow_origins=['*'], allow_methods=['*'], allow_headers=['*'])

@app.middleware('http')
async def production_cache_headers(request: Request, call_next):
    """Keep localhost and Vercel on the same frontend build.

    The UI changes frequently while balloon placement is being tuned, so stale JS/CSS
    must never survive a production deployment. Images/previews may still be cached by
    the browser for the life of the page.
    """
    response = await call_next(request)
    if request.url.path == '/' or request.url.path.endswith('.js') or request.url.path.endswith('.css'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# Local convenience: the Python backend also serves the plain HTML frontend.
# This makes one command enough: npm run dev OR python -m uvicorn backend.app.main:app --reload
if os.path.isdir(FRONTEND_ASSETS):
    app.mount('/assets', StaticFiles(directory=FRONTEND_ASSETS), name='frontend-assets')

@app.get('/')
def frontend_home():
    index_path = os.path.join(FRONTEND_DIR, 'index.html')
    if os.path.exists(index_path):
        return FileResponse(index_path, media_type='text/html')
    return {'ok': True, 'service': 'InspectBalloon API'}

@app.get('/favicon.ico')
def frontend_favicon():
    icon = os.path.join(FRONTEND_ASSETS, 'img', 'dfab-favicon.png')
    if os.path.exists(icon):
        return FileResponse(icon, media_type='image/png')
    raise HTTPException(404, 'Favicon not found')


class Balloon(BaseModel):
    number: int
    x: float
    y: float
    target_x: Optional[float] = None
    target_y: Optional[float] = None
    text: str = ''
    type: str = 'DIM'
    confidence: Optional[float] = None
    source: str = ''
    description: str = ''




def _renumber_balloons(items):
    """Return deep-copied balloons numbered continuously from 1..N."""
    result=[]
    for index, item in enumerate(items or [], start=1):
        clone=item.model_copy(deep=True) if hasattr(item, 'model_copy') else copy.deepcopy(item)
        clone.number=index
        result.append(clone)
    return result

def _renumber_rows(items):
    result=[]
    for index, item in enumerate(items or [], start=1):
        clone=item.model_copy(deep=True) if hasattr(item, 'model_copy') else copy.deepcopy(item)
        clone.number=index
        result.append(clone)
    return result

class LearnPayload(BaseModel):
    original_balloons: List[Balloon]
    final_balloons: List[Balloon]
    project_name: str = ''
    drawing_number: str = ''




class BatchAnalyzePayload(BaseModel):
    drawing_ids: List[str]
    force: bool = False

class DrawingExportPayload(BaseModel):
    balloons: List[Balloon]
    original_balloons: List[Balloon] = []
    learn: bool = False
    project_name: str = ''
    drawing_number: str = ''


class ProjectDrawingExport(BaseModel):
    drawing_id: str
    balloons: List[Balloon]
    original_balloons: List[Balloon] = []
    drawing_number: str = ''


class ProjectExportPayload(BaseModel):
    drawings: List[ProjectDrawingExport]
    learn: bool = False
    project_name: str = ''


class InspectionRow(BaseModel):
    number: int
    description: str = ''
    dimension: str = ''
    instrument: str = ''
    # Dynamic readings: one editable result column per inspected item.
    readings: List[str] = []
    # Backward compatibility with projects saved by earlier builds.
    reading1: str = ''
    reading2: str = ''


class InspectionReportPayload(BaseModel):
    drawing_id: str
    drawing_number: str = ''
    part_name: str = ''
    customer: str = ''
    revision: str = ''
    report_date: str = ''
    inspected_qty: str = ''
    total_qty: str = ''
    remarks: str = ''
    inspected_by: str = ''
    qc_incharge: str = ''
    approved_by: str = ''
    rows: List[InspectionRow] = []


class InspectionProjectPayload(BaseModel):
    project_name: str = ''
    reports: List[InspectionReportPayload]


class FinalPackagePayload(BaseModel):
    project_name: str = ''
    learn: bool = True
    drawings: List[ProjectDrawingExport]
    reports: List[InspectionReportPayload]


def _project_manifest_path(project_id: str) -> str:
    return os.path.join(PROJECTS, f'{project_id}.json')


def _read_manifest(project_id: str) -> dict:
    path = _project_manifest_path(project_id)
    if not os.path.exists(path):
        raise HTTPException(404, 'Project not found')
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _write_manifest(manifest: dict) -> None:
    with open(_project_manifest_path(manifest['project_id']), 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)


def _drawing_meta(drawing_id: str) -> dict:
    path = os.path.join(DATA, drawing_id, 'meta.json')
    if not os.path.exists(path):
        raise HTTPException(404, 'Drawing not found')
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _safe_name(value: str, fallback: str) -> str:
    cleaned = ''.join(c if c.isalnum() or c in ('-', '_', '.', ' ') else '_' for c in (value or '')).strip()
    return cleaned[:120] or fallback


def _save_learning(drawing_id: str, payload: LearnPayload) -> dict:
    originals = [b.model_dump() for b in payload.original_balloons]
    finals = [b.model_dump() for b in payload.final_balloons]

    def key(b):
        return (str(b.get('text', '')).strip().upper(), round(float(b.get('target_x') or 0) / 24), round(float(b.get('target_y') or 0) / 24))

    original_map = {key(b): b for b in originals}
    final_map = {key(b): b for b in finals}
    removed = [b.get('text', '') for k, b in original_map.items() if k not in final_map and b.get('text')]
    added = [b.get('text', '') for k, b in final_map.items() if k not in original_map and b.get('text')]

    # Match nearby balloon numbers for text edits where possible.
    edited = []
    by_number_original = {int(b.get('number', 0)): b for b in originals}
    by_number_final = {int(b.get('number', 0)): b for b in finals}
    for number, before in by_number_original.items():
        after = by_number_final.get(number)
        if after and before.get('text', '').strip() != after.get('text', '').strip():
            edited.append(f"{before.get('text', '')} -> {after.get('text', '')}")

    record = {
        'timestamp': datetime.now(timezone.utc).isoformat(),
        'drawing_id': drawing_id,
        'project_name': payload.project_name,
        'drawing_number': payload.drawing_number,
        'original_count': len(originals),
        'final_count': len(finals),
        'removed_texts': removed[:40],
        'added_texts': added[:40],
        'edited_texts': edited[:40],
    }
    path = os.path.join(LEARNING, 'corrections.jsonl')
    with open(path, 'a', encoding='utf-8') as f:
        f.write(json.dumps(record, ensure_ascii=False) + '\n')
    return record




def _analysis_cache_dir() -> str:
    path = os.path.join(DATA, 'analysis_cache')
    os.makedirs(path, exist_ok=True)
    return path


def _preview_hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def _cached_analysis(drawing_id: str):
    folder = os.path.join(DATA, drawing_id)
    local = os.path.join(folder, 'analysis.json')
    if os.path.exists(local):
        try:
            with open(local, 'r', encoding='utf-8') as f:
                result = json.load(f)
            result['cached'] = True
            return result
        except Exception:
            pass
    preview = os.path.join(folder, 'preview.png')
    if not os.path.exists(preview):
        return None
    key = _preview_hash(preview)
    shared = os.path.join(_analysis_cache_dir(), key + '.json')
    if os.path.exists(shared):
        try:
            with open(shared, 'r', encoding='utf-8') as f:
                result = json.load(f)
            result = {**result, 'drawing_id': drawing_id, 'cached': True}
            with open(local, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            return result
        except Exception:
            return None
    return None


def _store_analysis(drawing_id: str, result: dict):
    folder = os.path.join(DATA, drawing_id)
    local = os.path.join(folder, 'analysis.json')
    with open(local, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    preview = os.path.join(folder, 'preview.png')
    if os.path.exists(preview):
        key = _preview_hash(preview)
        shared = os.path.join(_analysis_cache_dir(), key + '.json')
        shared_result = dict(result)
        shared_result.pop('drawing_id', None)
        with open(shared, 'w', encoding='utf-8') as f:
            json.dump(shared_result, f, indent=2, ensure_ascii=False)


def _finalize_analysis_result(drawing_id: str, analysis: dict, cached: bool = False):
    characteristics = analysis.get('characteristics', [])
    if not characteristics:
        raise HTTPException(422, 'No inspectable characteristics were returned for this drawing.')
    meta = _drawing_meta(drawing_id)
    if analysis.get('drawing_number'):
        meta['drawing_number'] = analysis['drawing_number']
        with open(os.path.join(DATA, drawing_id, 'meta.json'), 'w', encoding='utf-8') as f:
            json.dump(meta, f, indent=2, ensure_ascii=False)
    result = {
        'drawing_id': drawing_id,
        'drawing_number': meta.get('drawing_number'),
        'part_name': analysis.get('part_name') or '',
        'customer': analysis.get('customer') or analysis.get('company_name') or '',
        'company_name': analysis.get('company_name') or '',
        'material': analysis.get('material') or '',
        'scale': analysis.get('scale') or '',
        'sheet_number': analysis.get('sheet_number') or '',
        'project_name': analysis.get('project_name') or '',
        'po_number': analysis.get('po_number') or '',
        'drawn_by': analysis.get('drawn_by') or '',
        'checked_by': analysis.get('checked_by') or '',
        'approved_by': analysis.get('approved_by') or '',
        'revision': analysis.get('revision') or '',
        'drawing_date': analysis.get('drawing_date') or '',
        'quantity': analysis.get('quantity') or '',
        'engine': 'analysis',
        'model': analysis.get('model'),
        'characteristics': characteristics,
        'count': len(characteristics),
        'cached': cached,
    }
    _store_analysis(drawing_id, result)
    return result


def _inspection_template_path() -> str:
    # One canonical template path for both localhost and Vercel.
    # Keeping it under backend/templates avoids duplicate api/assets copies.
    return os.path.join(BASE, 'templates', 'final_inspection_template.xlsx')


def _inspection_quantity(value: str) -> int:
    """Return the number of sample/result columns requested by INSPECTED QTY."""
    match = re.search(r'\d+', str(value or ''))
    if not match:
        return 2
    return max(1, min(int(match.group()), 50))


def _row_readings(item: InspectionRow, qty: int) -> list[str]:
    values = list(item.readings or [])
    if not values:
        values = [item.reading1 or '', item.reading2 or '']
    if len(values) < qty:
        values.extend([''] * (qty - len(values)))
    return values[:qty]


def _copy_cell_style(src, dst) -> None:
    """Copy template cell styling without changing merged/signature locations."""
    try:
        dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
    except Exception:
        pass


def _write_inspection_sheet(ws, report: InspectionReportPayload, rows: list[InspectionRow], page_no: int, total_pages: int) -> None:
    # Preserve the uploaded workbook's layout, merges, images/signatures and styling.
    ws['C7'] = report.part_name or ''
    ws['C8'] = report.drawing_number or ''
    ws['C9'] = report.customer or ''

    try:
        ws['J7'] = datetime.strptime(report.report_date, '%Y-%m-%d') if report.report_date else datetime.now()
    except ValueError:
        ws['J7'] = report.report_date or datetime.now()
    ws['J8'] = report.inspected_qty or ''
    ws['J9'] = report.total_qty or ''
    ws['M2'] = report.revision or '00'
    ws['M4'] = page_no

    qty = _inspection_quantity(report.inspected_qty)
    first_reading_col = 5  # E
    last_reading_col = first_reading_col + qty - 1

    # Dynamic sample headers: INSPECTED QTY 10 => result columns 1..10 (E:N).
    # We DO NOT insert worksheet columns.  This is intentional: the original
    # signature/approval cells and embedded signatures remain at their exact anchors.
    for col in range(first_reading_col, max(last_reading_col, 6) + 1):
        header = ws.cell(row=11, column=col)
        if col > 6:
            _copy_cell_style(ws.cell(row=11, column=6), header)
            # Give dynamically-created result cells the same practical width as F.
            letter = header.column_letter
            if ws.column_dimensions[letter].width is None:
                ws.column_dimensions[letter].width = ws.column_dimensions['F'].width or 14.4
        header.value = (col - first_reading_col + 1) if col <= last_reading_col else None

    # Clear characteristic rows while retaining the template formatting.
    clear_to = max(last_reading_col, 6)
    for row_no in range(12, 30):
        for col in range(1, clear_to + 1):
            cell = ws.cell(row=row_no, column=col)
            if col > 6:
                _copy_cell_style(ws.cell(row=row_no, column=6), cell)
            cell.value = None

    for offset, item in enumerate(rows[:18]):
        row_no = 12 + offset
        ws.cell(row=row_no, column=1).value = item.number
        ws.cell(row=row_no, column=2).value = item.description
        ws.cell(row=row_no, column=3).value = item.dimension
        ws.cell(row=row_no, column=4).value = item.instrument
        readings = _row_readings(item, qty)
        for idx, value in enumerate(readings):
            cell = ws.cell(row=row_no, column=first_reading_col + idx)
            if first_reading_col + idx > 6:
                _copy_cell_style(ws.cell(row=row_no, column=6), cell)
            cell.value = value or None

    ws['A32'] = 'Remarks: ' + (report.remarks or '')

    # Fixed large signature/stamp row for every generated template.
    # Row 34 holds names/signatures/stamps; row 35 keeps the role labels.
    ws.row_dimensions[34].height = max(float(ws.row_dimensions[34].height or 0), 72.0)
    for cell_ref in ('A34', 'E34', 'J34'):
        try:
            alignment = copy.copy(ws[cell_ref].alignment)
            alignment.horizontal = 'center'
            alignment.vertical = 'center'
            ws[cell_ref].alignment = alignment
        except Exception:
            pass

    # Keep the three approval/signature blocks at the EXACT original template anchors.
    # Never insert rows/columns before these cells.
    ws['A34'] = report.inspected_by or ''       # A34:D34
    ws['E34'] = report.qc_incharge or ''        # E34:I34
    ws['J34'] = report.approved_by or ''        # J34:M34


def _build_inspection_workbook(report: InspectionReportPayload, out_path: str) -> None:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(500, f'Excel export dependency is unavailable: {exc}')
    template = _inspection_template_path()
    if not os.path.exists(template):
        raise HTTPException(500, 'Inspection report template is missing')
    wb = load_workbook(template)
    source = wb[wb.sheetnames[0]]

    # Snapshot the uploaded template images before duplicating sheets. openpyxl does not
    # copy worksheet images, so create a fresh image object for every report page.
    from io import BytesIO
    from openpyxl.drawing.image import Image as XLImage
    image_specs = []
    for image in list(getattr(source, '_images', [])):
        try:
            image_specs.append({
                'data': image._data(),
                'anchor': copy.deepcopy(image.anchor),
                'width': image.width,
                'height': image.height,
            })
        except Exception:
            pass
    source._images = []

    def add_template_images(ws):
        # Signature/stamp images are re-anchored into the tall row directly above
        # INSPECTED BY / QC INCHARGE / APPROVED BY.  Using a fixed one-cell anchor
        # prevents Excel from stretching or shifting the stamps when row heights or
        # inspected-quantity columns change.
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        from openpyxl.utils import get_column_letter

        def col_width_px(col_idx: int) -> int:
            letter = get_column_letter(col_idx)
            width = ws.column_dimensions[letter].width
            width = 8.43 if width is None else float(width)
            return max(8, int(width * 7 + 5))

        signature_blocks = {
            'inspected': (1, 4),   # A:D
            'qc': (5, 9),          # E:I
            'approved': (10, 13),  # J:M
        }

        for spec in image_specs:
            try:
                original = copy.deepcopy(spec['anchor'])
                start = getattr(original, '_from', None)
                if start is None:
                    continue

                # Only approval-area stamp/signature images are repositioned.
                # Other images/logos retain the source workbook's original anchor.
                src_row = int(start.row)
                src_col = int(start.col)
                if not (28 <= src_row <= 34):
                    img = XLImage(BytesIO(spec['data']))
                    img.width = spec['width']; img.height = spec['height']
                    img.anchor = original
                    ws.add_image(img)
                    continue

                # Map the template's original stamp to its correct approval block.
                # Existing template anchors at H/I are QC; L/M are APPROVED.
                if src_col <= 3:
                    block = signature_blocks['inspected']
                elif src_col <= 8:
                    block = signature_blocks['qc']
                else:
                    block = signature_blocks['approved']

                start_col, end_col = block
                block_width = sum(col_width_px(c) for c in range(start_col, end_col + 1))
                row_height_pt = max(float(ws.row_dimensions[34].height or 72.0), 72.0)
                row_height_px = int(row_height_pt * 96 / 72)

                # Fit the stamp entirely inside the large signature row with margin,
                # preserve aspect ratio, and center it horizontally/vertically.
                aspect = (float(spec['width']) / float(spec['height'])) if spec['height'] else 1.0
                max_h = max(40, row_height_px - 12)
                max_w = max(55, block_width - 16)
                target_h = min(max_h, int(max_w / max(aspect, 0.01)))
                target_w = int(target_h * aspect)
                if target_w > max_w:
                    target_w = max_w
                    target_h = int(target_w / max(aspect, 0.01))

                x_off = max(4, (block_width - target_w) // 2)
                y_off = max(4, (row_height_px - target_h) // 2)
                marker = AnchorMarker(
                    col=start_col - 1,
                    colOff=pixels_to_EMU(x_off),
                    row=33,
                    rowOff=pixels_to_EMU(y_off),
                )
                extent = XDRPositiveSize2D(
                    cx=pixels_to_EMU(target_w),
                    cy=pixels_to_EMU(target_h),
                )

                img = XLImage(BytesIO(spec['data']))
                img.width = target_w
                img.height = target_h
                img.anchor = OneCellAnchor(_from=marker, ext=extent)
                ws.add_image(img)
            except Exception:
                # Never allow a bad stamp anchor to break report generation.
                pass

    add_template_images(source)
    rows = sorted(report.rows, key=lambda r: int(r.number))
    chunks = [rows[i:i + 18] for i in range(0, len(rows), 18)] or [[]]
    sheets = [source]
    for _ in range(1, len(chunks)):
        copied = wb.copy_worksheet(source)
        copied._images = []
        add_template_images(copied)
        sheets.append(copied)
    total_pages = len(chunks)
    for i, (ws, chunk) in enumerate(zip(sheets, chunks), start=1):
        ws.title = 'Inspection' if total_pages == 1 else f'Inspection {i}'
        _write_inspection_sheet(ws, report, chunk, i, total_pages)
    wb.save(out_path)


def _inspection_filename(report: InspectionReportPayload, fallback: str) -> str:
    return _safe_name(report.drawing_number, fallback) + '_inspection_report.xlsx'


@app.get('/health')
def health():
    learning_path = os.path.join(LEARNING, 'corrections.jsonl')
    learned = 0
    if os.path.exists(learning_path):
        try:
            with open(learning_path, 'r', encoding='utf-8') as f:
                learned = sum(1 for line in f if line.strip())
        except Exception:
            pass
    return {
        'ok': True,
        'service': 'InspectBalloon API',
        'analysis_configured': bool(os.getenv('GEMINI_API_KEY') or os.getenv('GEMINI_API_KEY_FALLBACK') or os.getenv('GEMINI_API_KEY_FALLBACK_2') or os.getenv('GOOGLE_API_KEY')),
        'analysis_key_slots': sum(bool((os.getenv(name) or '').strip()) for name in ('GEMINI_API_KEY','GEMINI_API_KEY_FALLBACK','GEMINI_API_KEY_FALLBACK_2','GOOGLE_API_KEY')),
        'analysis_model': os.getenv('GEMINI_MODEL', 'gemini-3.5-flash-lite'),
        'analysis_model_configured': bool(os.getenv('GEMINI_MODEL', '')),
        'learning_examples': learned,
        'runtime_storage': 'temporary' if (os.getenv('VERCEL') or DATA.startswith(tempfile.gettempdir())) else 'local',
        'fast_analysis': True,
    }


@app.post('/upload-batch')
async def upload_batch(files: List[UploadFile] = File(...), project_name: str = Form('')):
    if not files:
        raise HTTPException(400, 'Select at least one drawing')
    project_id = str(uuid.uuid4())
    manifest = {
        'project_id': project_id,
        'project_name': project_name or 'Drawing Ballooning Project',
        'drawings': [],
    }

    for file_index, file in enumerate(files):
        ext = os.path.splitext(file.filename or '')[1].lower()
        if ext not in ['.pdf', '.png', '.jpg', '.jpeg']:
            continue
        temp_id = str(uuid.uuid4())
        temp_folder = os.path.join(DATA, temp_id)
        os.makedirs(temp_folder, exist_ok=True)
        temp_src = os.path.join(temp_folder, 'source' + ext)
        with open(temp_src, 'wb') as f:
            shutil.copyfileobj(file.file, f)

        try:
            count = page_count(temp_src)
            for page_index in range(count):
                drawing_id = temp_id if page_index == 0 else str(uuid.uuid4())
                folder = os.path.join(DATA, drawing_id)
                if page_index > 0:
                    os.makedirs(folder, exist_ok=True)
                    shutil.copy2(temp_src, os.path.join(folder, 'source' + ext))
                src = os.path.join(folder, 'source' + ext)
                preview = os.path.join(folder, 'preview.png')
                render_page(src, preview, page_index)
                stem = os.path.splitext(os.path.basename(file.filename or f'drawing_{file_index+1}'))[0]
                provisional = stem if count == 1 else f'{stem} · Sheet {page_index+1}'
                meta = {
                    'drawing_id': drawing_id,
                    'project_id': project_id,
                    'source_filename': file.filename or f'drawing_{file_index+1}{ext}',
                    'page_index': page_index,
                    'page_number': page_index + 1,
                    'source_pages': count,
                    'drawing_number': provisional,
                }
                with open(os.path.join(folder, 'meta.json'), 'w', encoding='utf-8') as mf:
                    json.dump(meta, mf, indent=2, ensure_ascii=False)
                manifest['drawings'].append({
                    **meta,
                    'preview_url': f'/preview/{drawing_id}',
                    'status': 'uploaded',
                })
        except Exception as exc:
            shutil.rmtree(temp_folder, ignore_errors=True)
            raise HTTPException(400, f'Could not process {file.filename}: {exc}')

    if not manifest['drawings']:
        raise HTTPException(400, 'No supported PDF, PNG, JPG or JPEG drawings were found')
    _write_manifest(manifest)
    return manifest


@app.post('/upload-analyze-batch')
async def upload_analyze_batch(files: List[UploadFile] = File(...), project_name: str = Form('')):
    """Upload and analyse the entire drawing set inside one server request.

    Vercel serverless instances do not share /tmp reliably. The previous browser flow
    uploaded once and then fired six separate /analyze-ai requests; those requests
    could land on different instances, which is why production could show only one
    analysed drawing while localhost analysed all six. Keeping upload, page rendering
    and AI analysis in the same invocation removes that split-state failure mode.
    """
    manifest = await upload_batch(files=files, project_name=project_name)
    items = manifest.get('drawings', [])
    workers = max(1, min(int(os.getenv('GEMINI_PARALLEL_WORKERS', '2')), 3, len(items) or 1))
    result_by_id = {}

    def analyze_one(item):
        drawing_id = item['drawing_id']
        try:
            result = analyze_ai(drawing_id, force=False)
            return drawing_id, {'ok': True, **result}
        except HTTPException as exc:
            return drawing_id, {'ok': False, 'drawing_id': drawing_id, 'detail': str(exc.detail), 'status_code': exc.status_code}
        except Exception as exc:
            return drawing_id, {'ok': False, 'drawing_id': drawing_id, 'detail': str(exc), 'status_code': 502}

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(analyze_one, item) for item in items]
        for future in as_completed(futures):
            drawing_id, result = future.result()
            result_by_id[drawing_id] = result

    ready = 0
    failed = 0
    for item in items:
        result = result_by_id.get(item['drawing_id'], {'ok': False, 'detail': 'Analysis did not complete'})
        item['analysis'] = result
        if result.get('ok'):
            item['status'] = 'analyzed'
            ready += 1
        else:
            item['status'] = 'error'
            item['error'] = result.get('detail', 'Analysis failed')
            failed += 1

    manifest['analysis_summary'] = {
        'total': len(items),
        'successful': ready,
        'failed': failed,
        'parallel_workers': workers,
    }
    return manifest


@app.post('/reanalyze-upload')
async def reanalyze_upload(
    file: UploadFile = File(...),
    page_index: int = Form(0),
):
    """Re-analyse one failed drawing directly from the browser's original file.

    This endpoint is deliberately stateless. On Vercel, a retry can land on a new
    serverless instance where the earlier /tmp drawing no longer exists. Re-uploading
    just the original source file makes the small Analyze button reliable in both
    localhost and production.
    """
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext not in ['.pdf', '.png', '.jpg', '.jpeg']:
        raise HTTPException(400, 'Select a PDF, PNG, JPG or JPEG drawing')
    if not (os.getenv('GEMINI_API_KEY') or os.getenv('GEMINI_API_KEY_FALLBACK') or os.getenv('GEMINI_API_KEY_FALLBACK_2') or os.getenv('GOOGLE_API_KEY')):
        raise HTTPException(503, 'Analysis service is not configured')

    tmp_dir = tempfile.mkdtemp(prefix='balloon_retry_')
    src = os.path.join(tmp_dir, 'source' + ext)
    preview = os.path.join(tmp_dir, 'preview.png')
    try:
        with open(src, 'wb') as out:
            shutil.copyfileobj(file.file, out)
        count = page_count(src)
        page_index = max(0, min(int(page_index), max(0, count - 1)))
        render_page(src, preview, page_index)
        try:
            return {'ok': True, **analyze_with_gemini(preview)}
        except Exception as exc:
            text = str(exc)
            if '429' in text or 'RESOURCE_EXHAUSTED' in text:
                raise HTTPException(429, 'All configured analysis API-key/model fallbacks are currently quota-limited. Try this drawing again later.')
            raise HTTPException(502, f'Analysis failed: {text}')
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@app.post('/analyze-ai/{drawing_id}')
def analyze_ai(drawing_id: str, force: bool = False):
    folder = os.path.join(DATA, drawing_id)
    preview = os.path.join(folder, 'preview.png')
    if not os.path.exists(preview):
        raise HTTPException(404, 'Drawing not found')
    if not force:
        cached = _cached_analysis(drawing_id)
        if cached:
            return cached
    if not (os.getenv('GEMINI_API_KEY') or os.getenv('GEMINI_API_KEY_FALLBACK') or os.getenv('GEMINI_API_KEY_FALLBACK_2') or os.getenv('GOOGLE_API_KEY')):
        raise HTTPException(503, 'Analysis service is not configured. Add the API key to backend/.env and restart the backend.')
    try:
        analysis = analyze_with_gemini(preview)
    except Exception as exc:
        text = str(exc)
        if '429' in text or 'RESOURCE_EXHAUSTED' in text:
            raise HTTPException(429, 'Analysis quota is temporarily exhausted. Existing cached drawings are still available. Try again after the provider quota window resets or billing/quota is increased.')
        raise HTTPException(502, f'Analysis failed: {text}')
    return _finalize_analysis_result(drawing_id, analysis)


@app.post('/analyze-batch')
def analyze_batch(payload: BatchAnalyzePayload):
    # Attempt every supplied drawing. Older builds truncated this list to four,
    # which could leave a six-drawing upload partially analysed.
    ids = [x for x in payload.drawing_ids if x][:24]
    if not ids:
        raise HTTPException(400, 'No drawings supplied')

    workers = max(1, min(int(os.getenv('GEMINI_PARALLEL_WORKERS', '3')), 4, len(ids)))
    result_by_id = {}

    def analyze_one(drawing_id: str):
        try:
            result = analyze_ai(drawing_id, force=payload.force)
            return {**result, 'ok': True, 'cached': bool(result.get('cached', False))}
        except HTTPException as exc:
            return {'drawing_id': drawing_id, 'ok': False, 'detail': str(exc.detail), 'status_code': exc.status_code}
        except Exception as exc:
            return {'drawing_id': drawing_id, 'ok': False, 'detail': str(exc), 'status_code': 502}

    # Independent requests stop one malformed/slow drawing from cancelling its neighbours.
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(analyze_one, drawing_id): drawing_id for drawing_id in ids}
        for future in as_completed(futures):
            drawing_id = futures[future]
            try:
                result_by_id[drawing_id] = future.result()
            except Exception as exc:
                result_by_id[drawing_id] = {'drawing_id': drawing_id, 'ok': False, 'detail': str(exc), 'status_code': 502}

    results = [result_by_id.get(drawing_id, {'drawing_id': drawing_id, 'ok': False, 'detail': 'Analysis did not complete'}) for drawing_id in ids]
    return {
        'results': results,
        'total': len(ids),
        'successful': sum(1 for r in results if r.get('ok')),
        'cached': sum(1 for r in results if r.get('cached')),
        'parallel_workers': workers,
    }


@app.post('/analyze-project/{project_id}')
def analyze_project(project_id: str):
    manifest = _read_manifest(project_id)
    items = list(manifest['drawings'])
    # Small bounded parallelism improves batch speed without hammering Gemini and causing avoidable 503/429 responses.
    workers = max(1, min(int(os.getenv('GEMINI_PARALLEL_WORKERS', '2')), 4, len(items) or 1))
    result_by_id = {}

    def analyze_one(drawing_id: str):
        try:
            return {**analyze_ai(drawing_id), 'ok': True}
        except HTTPException as exc:
            return {'drawing_id': drawing_id, 'ok': False, 'detail': str(exc.detail)}
        except Exception as exc:
            return {'drawing_id': drawing_id, 'ok': False, 'detail': str(exc)}

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(analyze_one, d['drawing_id']): d['drawing_id'] for d in items}
        for future in as_completed(futures):
            drawing_id = futures[future]
            try:
                result_by_id[drawing_id] = future.result()
            except Exception as exc:
                result_by_id[drawing_id] = {'drawing_id': drawing_id, 'ok': False, 'detail': str(exc)}

    results = []
    for drawing in manifest['drawings']:
        result = result_by_id.get(drawing['drawing_id'], {'drawing_id': drawing['drawing_id'], 'ok': False, 'detail': 'Analysis did not complete'})
        if result.get('ok'):
            drawing['status'] = 'analyzed'
            drawing['drawing_number'] = result.get('drawing_number') or drawing.get('drawing_number')
            drawing['count'] = result.get('count', 0)
            drawing['analysis'] = result
        else:
            drawing['status'] = 'error'
            drawing['error'] = str(result.get('detail', 'Analysis failed'))
        results.append(result)

    _write_manifest(manifest)
    successful = sum(1 for r in results if r.get('ok'))
    if not successful:
        first_error = next((r.get('detail') for r in results if r.get('detail')), 'Analysis failed')
        raise HTTPException(502, first_error)
    return {'project_id': project_id, 'drawings': manifest['drawings'], 'results': results, 'successful': successful, 'total': len(results), 'parallel_workers': workers}


@app.get('/project/{project_id}')
def project(project_id: str):
    return _read_manifest(project_id)


@app.get('/preview/{drawing_id}')
def preview(drawing_id: str):
    path = os.path.join(DATA, drawing_id, 'preview.png')
    if not os.path.exists(path):
        raise HTTPException(404, 'Drawing not found')
    return FileResponse(path, media_type='image/png')


@app.post('/learn/{drawing_id}')
def learn(drawing_id: str, payload: LearnPayload):
    _drawing_meta(drawing_id)
    record = _save_learning(drawing_id, payload)
    return {'ok': True, 'saved': record}


@app.post('/inspection-report/{drawing_id}')
def inspection_report(drawing_id: str, payload: InspectionReportPayload):
    if payload.drawing_id != drawing_id:
        raise HTTPException(400, 'Drawing ID mismatch')
    _drawing_meta(drawing_id)
    out_dir = os.path.join(DATA, 'inspection_reports')
    os.makedirs(out_dir, exist_ok=True)
    filename = _inspection_filename(payload, 'drawing')
    out = os.path.join(out_dir, f'{drawing_id}_{filename}')
    _build_inspection_workbook(payload, out)
    return FileResponse(out, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=filename)


@app.post('/inspection-reports/{project_id}')
def inspection_reports(project_id: str, payload: InspectionProjectPayload):
    _read_manifest(project_id)
    if not payload.reports:
        raise HTTPException(400, 'No inspection reports supplied')
    out_dir = os.path.join(DATA, f'{project_id}_inspection_reports')
    shutil.rmtree(out_dir, ignore_errors=True)
    os.makedirs(out_dir, exist_ok=True)
    zip_path = os.path.join(DATA, f'{project_id}_inspection_reports.zip')
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for index, report in enumerate(payload.reports, start=1):
            _drawing_meta(report.drawing_id)
            filename = _inspection_filename(report, f'Drawing_{index:02d}')
            out = os.path.join(out_dir, filename)
            _build_inspection_workbook(report, out)
            zf.write(out, arcname=filename)
    filename = _safe_name(payload.project_name, 'inspection_reports') + '_inspection_reports.zip'
    return FileResponse(zip_path, media_type='application/zip', filename=filename)


@app.post('/final-package/{project_id}')
def final_package(project_id: str, payload: FinalPackagePayload):
    manifest = _read_manifest(project_id)
    if not payload.drawings or not payload.reports:
        raise HTTPException(400, 'Reviewed drawing data and inspection reports are required')

    drawing_by_id = {d.drawing_id: d for d in payload.drawings}
    report_by_id = {r.drawing_id: r for r in payload.reports}
    package_dir = os.path.join(DATA, f'{project_id}_final_package')
    shutil.rmtree(package_dir, ignore_errors=True)
    os.makedirs(package_dir, exist_ok=True)
    zip_path = os.path.join(DATA, f'{project_id}_final_package.zip')

    exported = 0
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for index, manifest_item in enumerate(manifest.get('drawings', []), start=1):
            drawing_id = manifest_item.get('drawing_id')
            drawing = drawing_by_id.get(drawing_id)
            report = report_by_id.get(drawing_id)
            if not drawing or not report:
                continue
            meta = _drawing_meta(drawing_id)
            folder = os.path.join(DATA, drawing_id)
            source_files = [x for x in os.listdir(folder) if x.startswith('source.')]
            if not source_files:
                continue

            display_number = report.drawing_number or drawing.drawing_number or meta.get('drawing_number') or f'Drawing_{index:02d}'
            safe = _safe_name(display_number, f'Drawing_{index:02d}')
            pair_dir = os.path.join(package_dir, safe)
            os.makedirs(pair_dir, exist_ok=True)

            pdf_name = safe + '_ballooned.pdf'
            pdf_path = os.path.join(pair_dir, pdf_name)
            src = os.path.join(folder, source_files[0])
            final_balloons = _renumber_balloons(drawing.balloons)
            export_ballooned_pdf(src, pdf_path, [b.model_dump() for b in final_balloons], int(meta.get('page_index', 0)))

            xlsx_name = safe + '_inspection_report.xlsx'
            xlsx_path = os.path.join(pair_dir, xlsx_name)
            _build_inspection_workbook(report, xlsx_path)

            zf.write(pdf_path, arcname=f'{safe}/{pdf_name}')
            zf.write(xlsx_path, arcname=f'{safe}/{xlsx_name}')
            exported += 1

            if payload.learn:
                _save_learning(drawing_id, LearnPayload(
                    original_balloons=drawing.original_balloons,
                    final_balloons=final_balloons,
                    project_name=payload.project_name,
                    drawing_number=display_number,
                ))

    if not exported:
        raise HTTPException(400, 'No reviewed drawing/report pairs were available to export')
    filename = _safe_name(payload.project_name, 'inspection_project') + '_final_package.zip'
    return FileResponse(zip_path, media_type='application/zip', filename=filename)


@app.post('/export-one/{drawing_id}')
def export_one(drawing_id: str, payload: DrawingExportPayload):
    meta = _drawing_meta(drawing_id)
    folder = os.path.join(DATA, drawing_id)
    files = [x for x in os.listdir(folder) if x.startswith('source.')]
    if not files:
        raise HTTPException(404, 'Source drawing missing')
    src = os.path.join(folder, files[0])
    out = os.path.join(folder, 'ballooned.pdf')
    final_balloons = _renumber_balloons(payload.balloons)
    export_ballooned_pdf(src, out, [b.model_dump() for b in final_balloons], int(meta.get('page_index', 0)))
    if payload.learn:
        _save_learning(drawing_id, LearnPayload(
            original_balloons=payload.original_balloons,
            final_balloons=final_balloons,
            project_name=payload.project_name,
            drawing_number=payload.drawing_number or meta.get('drawing_number', ''),
        ))
    name = _safe_name(payload.drawing_number or meta.get('drawing_number', ''), f'drawing_{meta.get("page_number", 1)}') + '_ballooned.pdf'
    return FileResponse(out, media_type='application/pdf', filename=name)


@app.post('/export-project/{project_id}')
def export_project(project_id: str, payload: ProjectExportPayload):
    manifest = _read_manifest(project_id)
    by_id = {d.drawing_id: d for d in payload.drawings}
    zip_path = os.path.join(DATA, f'{project_id}_ballooned.zip')
    export_dir = os.path.join(DATA, f'{project_id}_exports')
    shutil.rmtree(export_dir, ignore_errors=True)
    os.makedirs(export_dir, exist_ok=True)

    exported = 0
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for index, drawing in enumerate(manifest['drawings'], start=1):
            drawing_id = drawing['drawing_id']
            item = by_id.get(drawing_id)
            if not item:
                continue
            meta = _drawing_meta(drawing_id)
            folder = os.path.join(DATA, drawing_id)
            files = [x for x in os.listdir(folder) if x.startswith('source.')]
            if not files:
                continue
            src = os.path.join(folder, files[0])
            display_number = item.drawing_number or meta.get('drawing_number') or f'Drawing_{index:02d}'
            filename = _safe_name(display_number, f'Drawing_{index:02d}') + '_ballooned.pdf'
            out = os.path.join(export_dir, filename)
            final_balloons = _renumber_balloons(item.balloons)
            export_ballooned_pdf(src, out, [b.model_dump() for b in final_balloons], int(meta.get('page_index', 0)))
            zf.write(out, arcname=filename)
            exported += 1
            if payload.learn:
                _save_learning(drawing_id, LearnPayload(
                    original_balloons=item.original_balloons,
                    final_balloons=final_balloons,
                    project_name=payload.project_name,
                    drawing_number=display_number,
                ))

    if not exported:
        raise HTTPException(400, 'No reviewed drawings were available to export')
    filename = _safe_name(payload.project_name, 'ballooned_drawings') + '.zip'
    return FileResponse(zip_path, media_type='application/zip', filename=filename)
